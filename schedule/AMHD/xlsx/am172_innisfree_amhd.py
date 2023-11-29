import glob
import os
import shutil
import sys
import openpyxl
from datetime import datetime
from os.path import dirname


class AM172(object):
    def __init__(self):
        PATH = dirname(dirname(dirname(__file__)))
        sys.path.append(PATH)
        self.ADAPTER_RETAILER = 'innisfree_amhd'
        self.ADAPTER_TOKEN = '9efbf2b6aa6337f497c9a060b0fc2e658f27d06af0b514de3d50ba552af7f00a'
        self.FOLDER = 'innisfree_amhd'
        self.FULL_PATH = f'/home/{self.FOLDER}'
        self.EXT = '*xlsx'
        self.ORDERS = {}

    def scan_file(self):
        try:
            files = glob.glob(f'{self.FULL_PATH}/{self.EXT}')
            return max(files, key=os.path.getmtime)
        except:
            return None

    def get_data(self):
        from pos_api.adapter import submit_error, submit_order
        DATA = self.scan_file()
        if not DATA:
            submit_error(self.ADAPTER_RETAILER, 'FILE NOT FOUND')
            return
        dataframe = openpyxl.load_workbook(DATA, data_only=True)
        sheet = dataframe[dataframe.sheetnames[0]]
        orders = {}
        nRows = sheet.max_row + 1
        for row in range(3, nRows - 1):
            try:
                code = sheet[row][1].value
                pur_date = f'{sheet[row][2].value} {sheet[row][3].value}'
                pur_date = datetime.strptime(pur_date, '%Y-%m-%d %H:%M:%S')
                pur_date = pur_date.strftime('%Y-%m-%d %H:%M:%S')
                E = float(sheet[row][4].value)
                F = float(sheet[row][5].value)
                P = float(sheet[row][15].value)
                total = E - F - P
                discount = F + P
                vat = float(sheet[row][7].value)
                cash = float(sheet[row][8].value)
                card = float(sheet[row][9].value)
                card += float(sheet[row][10].value)
                card += float(sheet[row][11].value)
                card += float(sheet[row][12].value)
                card += float(sheet[row][13].value)
                card += float(sheet[row][14].value)
                voucher = float(sheet[row][17].value)
                gift_card = float(sheet[row][18].value)
                moca = float(sheet[row][19].value)
                now = float(sheet[row][20].value)
                air_pay = float(sheet[row][21].value)
                got_it = float(sheet[row][22].value)
                vn_pay = float(sheet[row][23].value)
                ur_box = float(sheet[row][24].value)
                gift_pop = float(sheet[row][25].value)
                cod = float(sheet[row][26].value)
                atome = float(sheet[row][27].value)
                momo = float(sheet[row][28].value)
                pms = [
                    {'Name': 'CASH', 'Value': cash},
                    {'Name': 'THẺ', 'Value': card},
                    {'Name': 'VOUCHER', 'Value': voucher},
                    {'Name': 'GIFT CARD', 'Value': gift_card},
                    {'Name': 'MOCA', 'Value': moca},
                    {'Name': 'NOW', 'Value': now},
                    {'Name': 'AIRPAY', 'Value': air_pay},
                    {'Name': 'GOT IT', 'Value': got_it},
                    {'Name': 'VNPAY', 'Value': vn_pay},
                    {'Name': 'URBOX', 'Value': ur_box},
                    {'Name': 'GIFT POP', 'Value': gift_pop},
                    {'Name': 'COD', 'Value': cod},
                    {'Name': 'ATOME', 'Value': atome},
                    {'Name': 'MOMO', 'Value': momo}
                ]
                for pm in pms.copy():
                    if not pm['Value']: pms.remove(pm)
                if not len(pms):
                    pms = [{'Name': 'CASH', 'Value': 0}]
                self.ORDERS.update({code: {
                    'Code': code,
                    'Status': 2,
                    'PurchaseDate': pur_date,
                    'Total': total,
                    'TotalPayment': total,
                    'VAT': vat,
                    'Discount': discount,
                    'PaymentMethods':pms,
                    'OrderDetails': [{'ProductId': 0}]
                }})

            except Exception as e:
                print(e)
                submit_error(retailer=self.ADAPTER_RETAILER, reason=str(e))
        for k, v in self.ORDERS.items():
            submit_order(retailer=self.ADAPTER_RETAILER, token=self.ADAPTER_TOKEN, data=v)
        self.backup(DATA)

    def backup(self, DATA):
        idx = DATA.rindex('/')
        name = DATA[idx + 1:]
        if os.path.exists(f'{self.FULL_PATH}/bak/{name}'):
            os.remove(f'{self.FULL_PATH}/bak/{name}')
        try:
            os.mkdir(f'{self.FULL_PATH}/bak')
        except:
            pass
        try:
            shutil.move(DATA, f'{self.FULL_PATH}/bak')
        except:
            pass
