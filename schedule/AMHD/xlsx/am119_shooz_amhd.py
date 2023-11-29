import glob
import os
import shutil
from datetime import datetime, timedelta
from os.path import dirname
import openpyxl


class AM119(object):
    def __init__(self):
        PATH = dirname(dirname(dirname(__file__)))
        sys.path.append(PATH)
        self.ADAPTER_RETAILER = 'shooz_amhd'
        self.ADAPTER_TOKEN = '0f01518005cdb6b3817e126de95f13d0599f3885a534d43795db3c047b04f70e'
        self.FOLDER = 'shooz_amhd'
        self.FULL_PATH = f'/home/{self.FOLDER}'
        self.EXT = '*xlsx'
        self.DATA = None
        self.ORDERS = {}
        self.PMS = {}
        self.ODS = {}

    def scan_file(self):
        try:
            files = glob.glob(f'{self.FULL_PATH}/{self.EXT}')
            return max(files, key=os.path.getmtime)
        except:
            return None

        # print(self.DATA)

    def get_data(self):
        from pos_api.adapter import submit_error, submit_order
        DATA = self.scan_file()
        if not DATA:
            submit_error(retailer=self.ADAPTER_RETAILER, reason='FILE_NOT_FOUND')
            return
        dataframe = openpyxl.load_workbook(DATA, data_only=True)
        raw = dataframe['Danh sách đơn hàng']
        orders = {}
        for row in range(2, raw.max_row + 1):
            try:
                pm = []
                code = raw[row][0].value
                if code is None: continue
                code = str(code).strip()
                code = code.replace("'", '').strip()
                if len(code) == 0: continue
                pur_date = raw[row][2].value
                pur_date = datetime.strptime(pur_date, '%d%m%Y%H%M')
                now = datetime.now() - timedelta(days=1)
                if pur_date.replace(hour=0, minute=0) < now.replace(hour=0, minute=0, second=0, microsecond=0): continue
                pur_date = pur_date.strftime('%Y-%m-%d %H:%M:%S')
                try:
                    total = float(raw[row][3].value)
                except:
                    total = 0
                vat = raw[row][4].value
                if orders.get(code) is None:
                    orders.update({code: {
                        'Code': code,
                        'Status': 2,
                        'PurchaseDate': pur_date,
                        'Total': total,
                        'TotalPayment': total,
                        'VAT': vat,
                        'Discount': 0,
                    }})
                else:
                    orders[code].update({
                        'Code': code,
                        'Status': 2,
                        # 'PurchaseDate': pur_date,
                        'Total': orders[code]['Total'] + total,
                        'TotalPayment': orders[code]['Total'] + total,
                        'VAT': orders[code]['VAT'] + vat,
                        'Discount': 0,
                    })
            except:
                pass
        raw = dataframe['Phương thức thanh toán']
        pms = {}
        for row in range(2, raw.max_row + 1):
            code = raw[row][0].value
            if code is None: continue
            code = str(code).strip()
            code = code.replace("'", '').strip()
            if len(code) == 0: continue
            name = str(raw[row][1].value).strip().upper()
            if name == 'NONE' or name == 'TIỀN MẶT': name = 'CASH'
            value = raw[row][2].value
            if pms.get(code) is None:
                pms.update({code: [{'Name': name, 'Value': value}]})
            else:
                pms[code].append({'Name': name, 'Value': value})
        for k, v in pms.items():
            if orders.get(k) is not None:
                orders[k].update({'PaymentMethods': v})
        raw = dataframe['Chi tiết đơn hàng']
        ods = {}
        for row in range(2, raw.max_row + 1):
            code = raw[row][0].value
            if code is None: continue
            code = str(code).strip()
            code = code.replace("'", '').strip()
            if len(code) == 0: continue
            p_code = raw[row][1].value
            p_code = p_code.replace("'", '')
            p_name = raw[row][2].value
            try:
                qty = float(raw[row][3].value)
            except:
                qty = 0
            try:
                price = raw[row][5].value
            except:
                price = 0
            if ods.get(code) is None:
                ods[code] = [{
                    'Code': p_code,
                    'Name': p_name,
                    'Price': price,
                    'Quantity': qty
                }]
            else:
                ods[code].append({
                    'Code': p_code,
                    'Name': p_name,
                    'Price': price,
                    'Quantity': qty
                })
        for k, v in ods.items():
            if orders.get(k) is not None:
                orders[k].update({'OrderDetails': v})
        for k, v in orders.items():
            if v.get('OrderDetails') is None:
                v.update({'OrderDetails': [{'ProductId': 0}]})
            if v.get('PaymentMethods') is None:
                v.update({'PaymentMethods': [{'Name': 'CASH', 'Value': 0}]})
            submit_order(retailer=self.ADAPTER_RETAILER, token=self.ADAPTER_TOKEN, data=v)
        self.backup(DATA)

    def backup(self, DATA):
        idx = DATA.rindex('/')
        name = DATA[idx + 1:]
        if os.path.exists(f'{self.FULL_PATH}/bak/{name}'):
            os.remove(f'{self.FULL_PATH}/bak/{name}')
        try:
            os.mkdir(f'{self.FULL_PATH}/bak')
        except:
            pass
        try:
            shutil.move(DATA, f'{self.FULL_PATH}/bak')
        except:
            pass
