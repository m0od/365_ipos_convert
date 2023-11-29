import glob
import os
import shutil
import sys
from datetime import datetime, timedelta
from os.path import dirname

import openpyxl


class AM164(object):
    def __init__(self):
        PATH = dirname(dirname(dirname(__file__)))
        sys.path.append(PATH)
        self.ADAPTER_RETAILER = 'alfresco_amhd'
        self.ADAPTER_TOKEN = '164fccd7b5172ec35763eb29c004491fb8281960311e90d06b9ee484db0e2f9e'
        self.FOLDER = 'alfresco_amhd'
        self.FULL_PATH = f'/home/{self.FOLDER}'
        self.DATE = datetime.now() - timedelta(days=1)
        self.EXT = f'*{self.DATE.strftime("%d.%m.%y")}*.xlsx'
        self.ORDERS = {}
        self.PMS = {}
        self.ODS = {}

    def scan_file(self):
        try:
            files = glob.glob(f'{self.FULL_PATH}/{self.EXT}')
            return max(files, key=os.path.getmtime)
        except:
            return None

    def get_data(self):
        from pos_api.adapter import submit_error, submit_order
        DATA = self.scan_file()
        if not DATA:
            submit_error(self.ADAPTER_RETAILER, 'FILE NOT FOUND')
            return
        dataframe = openpyxl.load_workbook(DATA, data_only=True)
        raw = dataframe['Danh sach don hang']
        orders = {}
        for row in range(10, raw.max_row + 1):
            # print(row)
            try:
                pm = []
                code = raw[row][5].value
                if code is None: break
                code = str(code).strip()
                if len(code) == 0: break
                pur_date = raw[row][7].value
                pur_date = datetime.strptime(pur_date, '%d%m%Y%H%M%S')
                now = datetime.now() - timedelta(days=1)
                pur_date = pur_date.strftime('%Y-%m-%d %H:%M:%S')
                try:
                    discount = abs(float(raw[row][8].value))
                except:
                    discount = 0
                try:
                    total = float(raw[row][9].value)
                except:
                    total = 0
                vat = raw[row][5].value
                if orders.get(code) is None:
                    orders.update({code: {
                        'Code': code,
                        'Status': 2,
                        'PurchaseDate': pur_date,
                        'Total': total,
                        'TotalPayment': total,
                        'VAT': 0,
                        'Discount': discount,
                    }})
                else:
                    orders[code].update({
                        'Code': code,
                        'Status': 2,
                        'PurchaseDate': pur_date,
                        'Total': orders[code]['Total'] + total,
                        'TotalPayment': orders[code]['Total'] + total,
                        'VAT': 0,
                        'Discount': orders[code]['Discount'] + discount,
                    })
            except:
                pass
        raw = dataframe['Phuong thuc thanh toan']
        pms = {}
        for row in range(10, raw.max_row + 1):
            code = raw[row][5].value
            # print(code)
            if code is None: break
            code = str(code).strip()
            if len(code) == 0: break
            name = str(raw[row][6].value).strip().upper()
            if name == 'NONE' or name == 'TIỀN MẶT': name = 'CASH'
            value = raw[row][7].value
            if pms.get(code) is None:
                pms.update({code: [{'Name': name, 'Value': value}]})
            else:
                pms[code].append({'Name': name, 'Value': value})
        for k, v in pms.items():
            if orders.get(k) is not None:
                orders[k].update({'PaymentMethods': v})
        for k, v in orders.items():
            if v.get('OrderDetails') is None:
                v.update({'OrderDetails': []})
            if v.get('PaymentMethods') is None:
                v.update({'PaymentMethods': [{'Name': 'CASH', 'Value': 0}]})
            submit_order(retailer=self.ADAPTER_RETAILER, token=self.ADAPTER_TOKEN, data=v)
        self.backup(DATA)

    def backup(self, DATA):
        idx = DATA.rindex('/')
        name = DATA[idx + 1:]
        if os.path.exists(f'{self.FULL_PATH}/bak/{name}'):
            os.remove(f'{self.FULL_PATH}/bak/{name}')
        try:
            os.mkdir(f'{self.FULL_PATH}/bak')
        except:
            pass
        try:
            shutil.move(DATA, f'{self.FULL_PATH}/bak')
        except:
            pass
