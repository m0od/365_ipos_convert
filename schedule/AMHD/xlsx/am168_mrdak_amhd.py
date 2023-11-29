import glob
import os
import shutil
from concurrent import futures
from datetime import datetime, timedelta
from os.path import dirname
import openpyxl
import sys

class AM168(object):
    def __init__(self):
        PATH = dirname(dirname(dirname(__file__)))
        sys.path.append(PATH)
        self.ADAPTER_RETAILER = 'mrdak_amhd'
        self.ADAPTER_TOKEN = '70f8a0a869f2ca7b25f9b68bdc86f374a2f541dc4e872a55118734c0c89bd112'
        self.FOLDER = 'mrdak_amhd'
        self.FULL_PATH = f'/home/{self.FOLDER}'
        self.method = {
            'TIỀN MẶT': 'CASH'
        }

    def scan_file(self, EXT):
        try:
            files = glob.glob(f'{self.FULL_PATH}/{EXT}')
            return max(files, key=os.path.getmtime)
        except:
            return None

    def get_data_hd(self):
        from pos_api.adapter import submit_error, submit_order
        DATA = self.scan_file('dakhd*xlsx')
        if not DATA:
            submit_error(self.ADAPTER_RETAILER, 'HD_NOT_FOUND')
            return
        try:
            dataframe = openpyxl.load_workbook(DATA, data_only=True)
            data = dataframe['Giao dịch']
            orders = {}
            for row in range(9, data.max_row + 1):
                pur_date = data[row][2].value
                if not pur_date: continue
                pur_date = datetime.strptime(pur_date.strip(), '%d/%m/%Y %H:%M')
                pur_date = pur_date.strftime('%Y-%m-%d %H:%M:%S')
                code = 'HD_' + str(data[row][3].value).strip()
                total = data[row][9].value
                discount = abs(data[row][12].value)
                vat = data[row][13].value
                pay_method = data[row][7].value.strip().upper()
                method = self.method.get(pay_method)
                method = method is not None and method or pay_method
                pay_total = data[row][14].value - data[row][16].value
                pm = [{'Name': method, 'Value': pay_total}]
                send = {
                    'Code': code,
                    'Status': 2,
                    'PurchaseDate': pur_date,
                    'Total': total,
                    'TotalPayment': total,
                    'VAT': vat,
                    'Discount': discount,
                    'OrderDetails': [],
                    'PaymentMethods': pm
                }
                submit_order(retailer=self.ADAPTER_RETAILER, token=self.ADAPTER_TOKEN, data=send)
        except Exception as e:
            submit_error(retailer=self.ADAPTER_RETAILER, reason=str(e))
            pass
        self.backup(DATA)

    def get_data_deli(self):
        from pos_api.adapter import submit_error, submit_order
        DATA = self.scan_file('dakdeli*xlsx')
        if not DATA:
            submit_error(self.ADAPTER_RETAILER, 'HD_NOT_FOUND')
            return
        try:
            dataframe = openpyxl.load_workbook(DATA, data_only=True)
            data = dataframe['Giao dịch']
            orders = {}
            for row in range(9, data.max_row + 1):
                pur_date = data[row][2].value
                if not pur_date: continue
                pur_date = datetime.strptime(pur_date.strip(), '%d/%m/%Y %H:%M')
                now = datetime.now() - timedelta(days=1)
                pur_date = pur_date.strftime('%Y-%m-%d %H:%M:%S')
                code = 'DELI_' + str(data[row][3].value).strip()
                total = data[row][7].value
                discount = abs(data[row][10].value)
                vat = data[row][11].value
                pay_method = data[row][5].value.strip().upper()
                method = self.method.get(pay_method)
                method = method is not None and method or pay_method
                pay_total = data[row][14].value - data[row][16].value
                pm = [{'Name': method, 'Value': pay_total}]
                send = {
                    'Code': code,
                    'Status': 2,
                    'PurchaseDate': pur_date,
                    'Total': total,
                    'TotalPayment': total,
                    'VAT': vat,
                    'Discount': discount,
                    'OrderDetails': [],
                    'PaymentMethods': pm
                }
                submit_order(retailer=self.ADAPTER_RETAILER, token=self.ADAPTER_TOKEN, data=send)
        except Exception as e:
            submit_error(self.ADAPTER_RETAILER, str(e))
            pass
        self.backup(DATA)

    def get_data(self):
        with futures.ThreadPoolExecutor(max_workers=2) as mt:
            thread = [
                mt.submit(self.get_data_hd),
                mt.submit(self.get_data_deli)
            ]
            futures.as_completed(thread)

    def backup(self, DATA):
        idx = DATA.rindex('/')
        name = DATA[idx + 1:]
        if os.path.exists(f'{self.FULL_PATH}/bak/{name}'):
            os.remove(f'{self.FULL_PATH}/bak/{name}')
        try:
            os.mkdir(f'{self.FULL_PATH}/bak')
        except:
            pass
        try:
            shutil.move(DATA, f'{self.FULL_PATH}/bak')
        except:
            pass

# MR_DAK_AMHD().get_data()
