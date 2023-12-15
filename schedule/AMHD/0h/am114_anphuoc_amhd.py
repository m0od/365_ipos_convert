import glob
import io
import os
import openpyxl
import shutil
import sys
from datetime import datetime, timedelta
from os.path import dirname


class AM114(object):
    def __init__(self):
        PATH = dirname(dirname(dirname(__file__)))
        sys.path.append(PATH)
        self.ADAPTER_RETAILER = 'anphuoc_amhd'
        self.ADAPTER_TOKEN = '3c8854e27a3694242ab157c9b404ad3c774900607a59b1264b5ae4f69bc69b7c'
        self.FOLDER = 'anphuoc_amhd'
        self.FULL_PATH = f'/home/{self.FOLDER}'
        self.DATE = datetime.now() - timedelta(days=1)
        self.EXT = f'*.xlsx'

    def scan_file(self):
        try:
            files = glob.glob(f'{self.FULL_PATH}/{self.EXT}')
            DATA = max(files, key=os.path.getmtime)
            t = os.path.getmtime(DATA)
            t = datetime.fromtimestamp(t).strftime('%Y%m%d %H:%M:%S')
            os.rename(DATA, f'{self.FULL_PATH}/{t}.xlsx')
            return f'{self.FULL_PATH}/{t}.xlsx'
        except:
            return None

    def backup(self, DATA):
        idx = DATA.rindex('/')
        name = DATA[idx + 1:]
        if os.path.exists(f'{self.FULL_PATH}/bak/{name}'):
            os.remove(f'{self.FULL_PATH}/bak/{name}')
        try:
            os.mkdir(f'{self.FULL_PATH}/bak')
        except:
            pass
        try:
            shutil.move(DATA, f'{self.FULL_PATH}/bak')
        except:
            pass

    def send_zero(self):
        from pos_api.adapter import submit_order
        order = {
            'Code': f'NOBILL_{self.DATE.strftime("%d%m%Y")}',
            'Status': 2,
            'PurchaseDate': self.DATE.strftime('%Y-%m-%d 07:00:00'),
            'Total': 0,
            'TotalPayment': 0,
            'VAT': 0,
            'Discount': 0,
            'OrderDetails': [{'ProductId': 0}],
            'PaymentMethods': [{'Name': 'CASH', 'Value': 0}]
        }
        submit_order(self.ADAPTER_RETAILER, self.ADAPTER_TOKEN, order)

    def get_data(self):
        from pos_api.adapter import submit_error, submit_order, submit_payment
        DATA = self.scan_file()
        if not DATA:
            submit_error(self.ADAPTER_RETAILER, 'FILE_NOT_FOUND')
            return
        dataframe = openpyxl.load_workbook(DATA, data_only=True)
        for sheet in dataframe.sheetnames:
            ws = dataframe[sheet]
            nRows = ws.max_row + 1
            orders = {}
            for row in range(1, nRows):
                code = ws[row][0].value
                if not code: continue
                if not len(str(code).strip()): continue
                code = str(code).strip()
                pur_date = str(ws[row][1].value).replace("'", '')[:12]
                if len(pur_date) < 12:
                    pur_date = pur_date.zfill(12)
                try:
                    pur_date = datetime.strptime(pur_date, '%d%m%Y%H%M')
                except:
                    continue
                if pur_date.strftime('%d%m%Y') != self.DATE.strftime('%d%m%Y'):
                    continue
                # if pur_date.strftime('%d%m%Y') != '20082023':
                #     continue
                discount = self.get_value(ws[row], 2)
                total = self.get_value(ws[row], 3)
                vat = self.get_value(ws[row], 4)
                cash = self.get_value(ws[row], 5)
                payoo = self.get_value(ws[row], 6)
                other = self.get_value(ws[row], 7)
                # print(code, pur_date, total, discount, vat, cash, payoo, other)
                if not orders.get(code):
                    orders.update({
                        code: {
                            'Code': code,
                            'Status': 2,
                            'PurchaseDate': pur_date.strftime('%Y-%m-%d %H:%M:%S'),
                            'Total': total,
                            'TotalPayment': total,
                            'VAT': vat,
                            'Discount': discount,
                            'OrderDetails': [{'ProductId': 0}],
                            'CASH': cash,
                            'PAYOO': payoo,
                            'OTHER': other
                        }
                    })
                else:
                    orders[code]['PurchaseDate'] = pur_date.strftime('%Y-%m-%d %H:%M:%S')
                    orders[code]['Total'] += total
                    orders[code]['TotalPayment'] += total
                    orders[code]['VAT'] += vat
                    orders[code]['Discount'] += discount
                    orders[code]['CASH'] += cash
                    orders[code]['PAYOO'] += payoo
                    orders[code]['OTHER'] += other
            for _, order in orders.items():
                pms = [
                    {'Name': 'CASH', 'Value': order['CASH']},
                    {'Name': 'PAYOO', 'Value': order['PAYOO']},
                    {'Name': 'OTHER', 'Value': order['OTHER']},
                ]
                for pm in pms.copy():
                    if pm['Value'] == 0:
                        pms.remove(pm)
                order.update({'PaymentMethods': pms})
                if not len(order['PaymentMethods']):
                    order['PaymentMethods'] = [{
                        'Name': 'CASH', 'Value': 0
                    }]
                order.pop('CASH')
                order.pop('PAYOO')
                order.pop('OTHER')
                submit_order(self.ADAPTER_RETAILER, self.ADAPTER_TOKEN, order)
                need_fix = False
                for pm in order['PaymentMethods']:
                    if pm['Value'] <= 0:
                        need_fix = True
                        break
                if need_fix:
                    for pm in order['PaymentMethods']:
                        pm = {
                            'Code': f'{order["Code"]}-{pm["Name"]}',
                            'OrderCode': order['Code'],
                            'Amount': pm['Value'],
                            'TransDate': order['PurchaseDate'],
                            'AccountId': pm["Name"]
                        }
                        submit_payment(self.ADAPTER_RETAILER, self.ADAPTER_TOKEN, pm)
        self.backup(DATA)

    def get_value(self, row, idx):
        try:
            value = str(row[idx].value).replace(',', '')
            value = float(value)
        except:
            value = 0
        return value
