import glob
import os
import random
import shutil
import sys
import openpyxl
from datetime import datetime, timedelta
from os.path import dirname


class AM197(object):
    def __init__(self):
        PATH = dirname(dirname(dirname(__file__)))
        sys.path.append(PATH)
        self.ADAPTER_RETAILER = 'ecco_amhd'
        self.ADAPTER_TOKEN = '021e1a5a0a00ae143f4794337d9a2cf0473f72e39af6b6616e70ff0ae7d14090'
        self.FOLDER = 'ecco_amhd'
        self.FULL_PATH = f'/home/{self.FOLDER}'
        self.DATE = datetime.now() - timedelta(days=1)
        self.EXT = f'*.XLSX'

    def scan_file(self):
        try:
            files = glob.glob(f'{self.FULL_PATH}/{self.EXT}')
            return max(files, key=os.path.getmtime)
        except:
            return None

    def backup(self, DATA):
        idx = DATA.rindex('/')
        name = DATA[idx + 1:]
        if os.path.exists(f'{self.FULL_PATH}/bak/{name}'):
            os.remove(f'{self.FULL_PATH}/bak/{name}')
        try:
            os.mkdir(f'{self.FULL_PATH}/bak')
        except:
            pass
        try:
            shutil.move(DATA, f'{self.FULL_PATH}/bak')
        except:
            pass

    def get_data(self):
        from pos_api.adapter import submit_error, submit_order
        DATA = self.scan_file()
        if not DATA:
            submit_error(self.ADAPTER_RETAILER, 'FILE_NOT_FOUND')
            return
        dataframe = openpyxl.load_workbook(DATA, data_only=True)
        sheet = dataframe[dataframe.sheetnames[0]]
        nRows = sheet.max_row + 1
        orders = []
        headers = []
        for _ in sheet[1]:
            headers.append(str(_.value).strip().upper())
        for row in range(2, nRows):
            rec = dict(zip(headers, sheet[row]))
            pur_date = f'{str(rec["DATE"].value)[:-8].strip()} {rec["TIME"].value}'
            try:
                # print(pur_date)
                pur_date = datetime.strptime(pur_date, '%Y-%m-%d %H:%M:%S')
            except:
                continue
            # print(pur_date)
            code = str(rec["BILL_NO"].value).strip()
            total = round(float(rec["TOTAL"].value), 0)
            pm = str(rec["TENDER"].value).strip()
            # print(sheet[row][2].value)
            # total = sheet[row][2].value
            # print(sheet[row][3].value)
            # code = f'HD-{pur_date.strftime("%d%m%y")}-{len(orders) + 1:05d}'
            orders.append({
                'Code': code,
                'Status': 2,
                'PurchaseDate': pur_date.strftime('%Y-%m-%d %H:%M:%S'),
                'Total': total,
                'TotalPayment': total,
                'VAT': 0,
                'Discount': 0,
                'OrderDetails': [{'ProductId': 0}],
                'PaymentMethods': [{'Name': pm, 'Value': total}]
            })
            # print(orders[-1])
        for order in orders:
            submit_order(self.ADAPTER_RETAILER, self.ADAPTER_TOKEN, order)
        self.backup(DATA)
