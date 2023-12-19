import glob
import os
import random
import shutil
import sys
import openpyxl
from datetime import datetime, timedelta
from os.path import dirname


class AM193(object):
    def __init__(self):
        PATH = dirname(dirname(dirname(__file__)))
        sys.path.append(PATH)
        self.ADAPTER_RETAILER = 'miniso_amhd'
        self.ADAPTER_TOKEN = '4da28336ead243b84d4f4e187f22903e3eec48a00cb61765c9b1d2df4bb784a9'
        self.FOLDER = 'miniso_amhd'
        self.FULL_PATH = f'/home/{self.FOLDER}'
        self.DATE = datetime.now() - timedelta(days=1)
        self.EXT = f'*.xlsx'

    def scan_file(self):
        try:
            files = glob.glob(f'{self.FULL_PATH}/{self.EXT}')
            return max(files, key=os.path.getmtime)
        except:
            return None

    def backup(self, DATA):
        idx = DATA.rindex('/')
        name = DATA[idx + 1:]
        if os.path.exists(f'{self.FULL_PATH}/bak/{name}'):
            os.remove(f'{self.FULL_PATH}/bak/{name}')
        try:
            os.mkdir(f'{self.FULL_PATH}/bak')
        except:
            pass
        try:
            shutil.move(DATA, f'{self.FULL_PATH}/bak')
        except:
            pass

    def send_zero(self):
        from pos_api.adapter import submit_order
        order = {
            'Code': f'NOBILL_{self.DATE.strftime("%d%m%Y")}',
            'Status': 2,
            'PurchaseDate': self.DATE.strftime('%Y-%m-%d 07:00:00'),
            'Total': 0,
            'TotalPayment': 0,
            'VAT': 0,
            'Discount': 0,
            'OrderDetails': [{'ProductId': 0}],
            'PaymentMethods': [{'Name': 'CASH', 'Value': 0}]
        }
        submit_order(self.ADAPTER_RETAILER, self.ADAPTER_TOKEN, order)

    def get_data(self):
        from pos_api.adapter import submit_error, submit_order
        DATA = self.scan_file()
        if not DATA:
            submit_error(self.ADAPTER_RETAILER, 'FILE_NOT_FOUND')
            return
        dataframe = openpyxl.load_workbook(DATA, data_only=True)
        sheet = dataframe[dataframe.sheetnames[0]]
        nRows = sheet.max_row + 1
        count = 0
        orders = []
        for row in range(3, nRows):
            pur_date = sheet[row][0].value
            try:
                pur_date = datetime.strptime(pur_date, '%Y-%m-%d')
            except:
                continue
            count += 1
            qty = float(sheet[row][24].value)
            price = float(sheet[row][31].value)
            ods = [{
                'Code': sheet[row][11].value,
                'Name': sheet[row][13].value,
                'Price': price / qty,
                'Quantity': qty

            }]
            code = f'HD-{pur_date.strftime("%d%m%y")}-{count:05d}'
            orders.append({
                'Code': code,
                'Status': 2,
                'PurchaseDate': pur_date,
                'Total': price,
                'TotalPayment': price,
                'VAT': 0,
                'Discount': 0,
                'OrderDetails': ods,
                'PaymentMethods': [{'Name': 'CASH', 'Value': price}]
            })
        time = []
        for idx, order in enumerate(orders):
            h = 10 + int(11 * idx/len(orders))
            pur_date = order['PurchaseDate'].strftime('%Y-%m-%d')
            m = random.randint(0, 59)
            s = random.randint(0, 59)
            pur_date = f'{pur_date} {h:02d}:{m:02d}:{s:02d}'
            time.append(pur_date)
        time.sort()
        for idx, t in enumerate(time):
            orders[idx].update({'PurchaseDate': t})
            submit_order(self.ADAPTER_RETAILER, self.ADAPTER_TOKEN, orders[idx])
        if not count:
            self.send_zero()
        self.backup(DATA)
