import glob
import os
import shutil
import sys
from datetime import datetime, timedelta
from os.path import dirname
import openpyxl


class AM142(object):
    def __init__(self):
        PATH = dirname(dirname(dirname(__file__)))
        sys.path.append(PATH)
        self.ADAPTER_RETAILER = 'routine_amhd'
        self.ADAPTER_TOKEN = '9016a35bbb92076a998c777bb7cf3924ec1c211692fdc76314d277b0f496a41b'
        self.FOLDER = 'routine_amhd'
        self.FULL_PATH = f'/home/{self.FOLDER}'
        self.DATE = datetime.now() - timedelta(days=1)
        self.EXT = f'*xlsx'
        self.ORDERS = {}
        self.PMS = {}
        self.ODS = {}

    def scan_file(self):
        try:
            files = glob.glob(f'{self.FULL_PATH}/{self.EXT}')
            return max(files, key=os.path.getmtime)
        except:
            return None

    def get_data(self):
        from pos_api.adapter import submit_error, submit_order
        DATA = self.scan_file()
        if not DATA:
            submit_error(self.ADAPTER_RETAILER, 'FILE_NOT_FOUND')
            return
        dataframe = openpyxl.load_workbook(DATA, data_only=True)
        raw = dataframe['Danh sách đơn hàng']
        for row in range(2, raw.max_row + 1):
            code = raw[row][0].value
            if code is None: continue
            code = code.strip()
            if len(code) == 0: continue
            pur_date = raw[row][2].value
            # if pur_date.strftime('%Y%m%d') != self.DATE.strftime('%Y%m%d'):
            #     continue
            pur_date = pur_date.strftime('%Y-%m-%d %H:%M:%S')
            if self.ORDERS.get(code) is None:
                self.ORDERS.update({str(code).strip(): {}})
            status = raw[row][1].value
            discount = raw[row][3].value
            discount = discount is not None and abs(discount) or 0
            if status == 2:
                total = raw[row][4].value
                vat = raw[row][5].value
            else:
                total = raw[row][4].value * -1
                vat = raw[row][5].value * -1
            total_pm = raw[row][6].value
            self.ORDERS.update({code: {
                'Code': str(code),
                'Status': 2,
                'PurchaseDate': pur_date,
                'Total': total,
                'TotalPayment': total_pm,
                'VAT': vat,
                'Discount': discount,
            }})
        raw = dataframe['Phương thức thanh toán']
        for row in range(2, raw.max_row + 1):
            code = raw[row][0].value
            if code is None: continue
            code = str(code).strip()
            if not len(code): continue
            name = str(raw[row][1].value).replace('TTTM', '').strip()
            value = raw[row][2].value
            if self.PMS.get(code) is None:
                self.PMS[code] = {
                    'PaymentMethods': [
                        {
                            'Name': name,
                            'Value': float(value)
                        }
                    ]
                }
            else:
                self.PMS[code]['PaymentMethods'].append(
                    {
                        'Name': name,
                        'Value': float(value)
                    }
                )
        for k, v in self.PMS.items():
            if self.ORDERS.get(k) is not None:
                self.ORDERS[k].update(v)
        raw = dataframe['Chi tiết đơn hàng']
        self.ODS = {}
        for row in range(2, raw.max_row + 1):
            code = raw[row][0].value
            if code is None: continue
            code = str(code).strip()
            p_code = str(raw[row][1].value)
            name = str(raw[row][2].value)
            qty = raw[row][3].value
            price = raw[row][5].value
            price = price is not None and price or 0
            if self.ODS.get(code) is None:
                self.ODS[code] = {
                    'OrderDetails': [{
                        'Code': p_code,
                        'Name': name,
                        'Price': float(price),
                        'Quantity': int(qty)
                    }],
                }
            else:
                self.ODS[code]['OrderDetails'].append(
                    {
                        'Code': p_code,
                        'Name': name,
                        'Price': float(price),
                        'Quantity': int(qty)
                    }
                )
        for k, v in self.ODS.items():
            if self.ORDERS.get(k) is not None:
                self.ORDERS[k].update(v)
        for k, v in self.ORDERS.items():
            if v.get('Code') is not None:
                if v.get('PaymentMethods') is None:
                    v['PaymentMethods'] = [{'Name': 'CASH', 'Value': 0}]
                if v.get('OrderDetails') is None:
                    v['OrderDetails'] = [{'ProductId': 0}]
                submit_order(retailer=self.ADAPTER_RETAILER, token=self.ADAPTER_TOKEN, data=v)
        self.backup(DATA)
        # if not len(list(self.ORDERS.keys())):
        #     self.send_zero()

    def backup(self, DATA):
        idx = DATA.rindex('/')
        name = DATA[idx + 1:]
        if os.path.exists(f'{self.FULL_PATH}/bak/{name}'):
            os.remove(f'{self.FULL_PATH}/bak/{name}')
        try:
            os.mkdir(f'{self.FULL_PATH}/bak')
        except:
            pass
        try:
            shutil.move(DATA, f'{self.FULL_PATH}/bak')
        except:
            pass

    def send_zero(self):
        from pos_api.adapter import submit_order
        order = {
            'Code': f'NOBILL_{self.DATE.strftime("%d%m%Y")}',
            'Status': 2,
            'PurchaseDate': self.DATE.strftime('%Y-%m-%d 07:00:00'),
            'Total': 0,
            'TotalPayment': 0,
            'VAT': 0,
            'Discount': 0,
            'OrderDetails': [{'ProductId': 0}],
            'PaymentMethods': [{'Name': 'CASH', 'Value': 0}]
        }
        submit_order(self.ADAPTER_RETAILER, self.ADAPTER_TOKEN, order)
