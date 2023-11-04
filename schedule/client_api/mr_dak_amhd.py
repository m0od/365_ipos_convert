import glob
import os
import shutil
from concurrent import futures
from datetime import datetime, timedelta
from os.path import dirname
import openpyxl


class MR_DAK_AMHD(object):
    def __init__(self):
        self.ADAPTER_RETAILER = 'mrdak_amhd'
        self.ADAPTER_TOKEN = '70f8a0a869f2ca7b25f9b68bdc86f374a2f541dc4e872a55118734c0c89bd112'
        # self.ADAPTER_RETAILER = '695'
        # self.ADAPTER_TOKEN = 'cf0f760c3c11b65139beaecd6e0dd12f80bc34a177704ffc497d2bf816d1ac2d'

        self.FOLDER = 'mrdak_amhd'
        self.FULL_PATH = f'../home/{self.FOLDER}/'
        self.HD = 'dakhd*xlsx'
        self.DELI = 'dakdeli*xlsx'
        # self.DATA = None
        self.method = {
            'TIỀN MẶT': 'CASH'
        }

    def scan_file(self, name):
        # print(os.getcwd())
        # print(os.path.getmtime(self.FULL_PATH))
        # print(for name in glob.glob('/home/geeks/Desktop/gfg/data.txt'):)
        if name == 'HD':
            files = glob.glob(self.FULL_PATH + self.HD)
        else:
            files = glob.glob(self.FULL_PATH + self.DELI)
        # print(files)
        return max(files, key=os.path.getmtime)

        # print(self.DATA)

    def get_data_hd(self):
        try:
            DATA = self.scan_file('HD')
            dataframe = openpyxl.load_workbook(DATA, data_only=True)
            data = dataframe['Giao dịch']
            orders = {}
            # Iterate the loop to read the cell values
            for row in range(9, data.max_row + 1):
                pur_date = data[row][2].value.strip()
                pur_date = datetime.strptime(pur_date, '%d/%m/%Y %H:%M')
                pur_date = pur_date.strftime('%Y-%m-%d %H:%M:%S')
                # print(pur_date)
                code = 'HD_' + str(data[row][3].value).strip()
                total = data[row][9].value
                discount = abs(data[row][12].value)
                vat = data[row][13].value
                pay_method = data[row][7].value.strip().upper()
                method = self.method.get(pay_method)
                method = method is not None and method or pay_method
                pay_total = data[row][14].value - data[row][16].value
                pm = [{'Name': method, 'Value': pay_total}]
                send = {
                    'Code': code,
                    'Status': 2,
                    'PurchaseDate': pur_date,
                    'Total': total,
                    'TotalPayment': total,
                    'VAT': vat,
                    'Discount': discount,
                    'OrderDetails': [],
                    'PaymentMethods': pm
                }
                # print(send)
                submit_order(retailer=self.ADAPTER_RETAILER, token=self.ADAPTER_TOKEN, data=send)

        except Exception as e:
            submit_error(retailer=self.ADAPTER_RETAILER, reason=str(e))
            pass
        idx = DATA.rindex('/')
        name = DATA[idx + 1:]
        if os.path.exists(f'{self.FULL_PATH}bak/{name}'):
            os.remove(f'{self.FULL_PATH}bak/{name}')
        try:
            shutil.move(DATA, f'{self.FULL_PATH}bak')
        except:
            pass

    def get_data_deli(self):
        try:
            DATA = self.scan_file('DELI')
            dataframe = openpyxl.load_workbook(DATA, data_only=True)
            data = dataframe['Giao dịch']
            orders = {}
            # Iterate the loop to read the cell values
            for row in range(9, data.max_row + 1):
                # print(data[row][2].value)
                pur_date = data[row][2].value.strip()
                pur_date = datetime.strptime(pur_date, '%d/%m/%Y %H:%M')
                now = datetime.now() - timedelta(days=1)
                # if pur_date.replace(hour=0, minute=0) < now.replace(hour=0, minute=0, second=0, microsecond=0): continue
                pur_date = pur_date.strftime('%Y-%m-%d %H:%M:%S')
                # print(pur_date)
                code = 'DELI_' + str(data[row][3].value).strip()
                total = data[row][7].value
                discount = abs(data[row][10].value)
                vat = data[row][11].value
                pay_method = data[row][5].value.strip().upper()
                method = self.method.get(pay_method)
                method = method is not None and method or pay_method
                pay_total = data[row][14].value - data[row][16].value
                pm = [{'Name': method, 'Value': pay_total}]
                send = {
                    'Code': code,
                    'Status': 2,
                    'PurchaseDate': pur_date,
                    'Total': total,
                    'TotalPayment': total,
                    'VAT': vat,
                    'Discount': discount,
                    'OrderDetails': [],
                    'PaymentMethods': pm
                }
                # print(send)
                submit_order(retailer=self.ADAPTER_RETAILER, token=self.ADAPTER_TOKEN, data=send)
        except Exception as e:
            submit_error(retailer=self.ADAPTER_RETAILER, reason=str(e))
            pass
        idx = DATA.rindex('/')
        name = DATA[idx + 1:]
        if os.path.exists(f'{self.FULL_PATH}bak/{name}'):
            os.remove(f'{self.FULL_PATH}bak/{name}')
        try:
            shutil.move(DATA, f'{self.FULL_PATH}bak')
        except:
            pass

    def get_data(self):
        with futures.ThreadPoolExecutor(max_workers=2) as mt:
            thread = [
                mt.submit(self.get_data_hd),
                mt.submit(self.get_data_deli)
            ]
            futures.as_completed(thread)


if __name__:
    import sys

    # PATH = dirname(dirname(__file__))
    PATH = dirname(dirname(dirname(__file__)))
    sys.path.append(PATH)
    from schedule.pos_api.adapter import submit_error, submit_order
    MR_DAK_AMHD().get_data()
