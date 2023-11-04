import glob
import os
import shutil
from datetime import datetime, timedelta
from os.path import dirname

import openpyxl



class ALFRESCO_AMHD(object):
    def __init__(self):
        self.ADAPTER_RETAILER = 'alfresco_amhd'
        self.ADAPTER_TOKEN = '164fccd7b5172ec35763eb29c004491fb8281960311e90d06b9ee484db0e2f9e'
        # self.ADAPTER_RETAILER = '695'
        # self.ADAPTER_TOKEN = 'cf0f760c3c11b65139beaecd6e0dd12f80bc34a177704ffc497d2bf816d1ac2d'

        self.FOLDER = 'alfresco_amhd'
        self.FULL_PATH = f'../home/{self.FOLDER}/'
        self.EXT = '*xlsx'
        self.DATA = None
        self.ORDERS = {}
        self.PMS = {}
        self.ODS = {}

    def scan_file(self):
        # print(os.getcwd())
        # print(os.path.getmtime(self.FULL_PATH))
        # print(for name in glob.glob('/home/geeks/Desktop/gfg/data.txt'):)
        files = glob.glob(self.FULL_PATH + self.EXT)
        # print(files)
        self.DATA = max(files, key=  os.path.getmtime)

        print(self.DATA)


    def get_data(self):
        self.scan_file()
        dataframe = openpyxl.load_workbook(self.DATA, data_only=True)

        sheet1 = dataframe['Danh sach don hang']
        orders = {}
        # Iterate the loop to read the cell values
        for row in range(10, sheet1.max_row + 1):
            # print(row)
            try:
                pm = []
                code = sheet1[row][5].value
                if code is None: break
                code = str(code).strip()
                if len(code) == 0: break
                pur_date = sheet1[row][7].value
                pur_date = datetime.strptime(pur_date, '%d%m%Y%H%M%S')
                now = datetime.now() - timedelta(days=1)
                # if pur_date.replace(hour=0, minute=0) < now.replace(hour=0, minute=0, second=0, microsecond=0): continue
                pur_date = pur_date.strftime('%Y-%m-%d %H:%M:%S')
                # print(code, pur_date)

                try:
                    discount = abs(float(sheet1[row][8].value))
                except:
                    discount = 0
                try:
                    total = float(sheet1[row][9].value)
                except:
                    total = 0
                # print(code, pur_date, discount, total)
                vat = sheet1[row][5].value
                if orders.get(code) is None:
                    orders.update({code:{
                        'Code': code,
                        'Status': 2,
                        'PurchaseDate': pur_date,
                        'Total': total,
                        'TotalPayment': total,
                        'VAT': 0,
                        'Discount': discount,
                    }})
                else:
                    orders[code].update({
                        'Code': code,
                        'Status': 2,
                        'PurchaseDate': pur_date,
                        'Total': orders[code]['Total'] + total,
                        'TotalPayment': orders[code]['Total'] + total,
                        'VAT': 0,
                        'Discount': orders[code]['Discount'] + discount,
                    })
            except Exception as e:
                print(e)
                # submit_error(retailer=self.ADAPTER_RETAILER, reason=str(e))
        # print('abcsdsda')
        sheet2 = dataframe['Phuong thuc thanh toan']
        pms = {}
        for row in range(10, sheet2.max_row + 1):
            code = sheet2[row][5].value
            # print(code)
            if code is None: break
            code = str(code).strip()
            if len(code) == 0: break
            name = str(sheet2[row][6].value).strip().upper()
            if name == 'NONE' or name == 'TIỀN MẶT': name = 'CASH'
            # elif len(name) == 0:
            value = sheet2[row][7].value
            if pms.get(code) is None:
                pms.update({code: [{'Name': name, 'Value': value}]})
            else:
                pms[code].append({'Name': name, 'Value': value})
        for k, v in pms.items():
            if orders.get(k) is not None:
                orders[k].update({'PaymentMethods': v})
        for k, v in orders.items():
            if v.get('OrderDetails') is None:
                v.update({'OrderDetails': []})
            if v.get('PaymentMethods') is None:
                v.update({'PaymentMethods': [{'Name': 'CASH', 'Value': 0}]})
            submit_order(retailer=self.ADAPTER_RETAILER, token=self.ADAPTER_TOKEN, data=v)
        try:
            shutil.move(self.DATA, f"{self.FULL_PATH}bak")
        except:
            pass
if __name__:
    # print(__name__)
    import sys

    PATH = dirname(dirname(__file__))
    # PATH = dirname(dirname(dirname(__file__)))
    # print(PATH)
    sys.path.append(PATH)
    from schedule.pos_api.adapter import submit_error, submit_order
    # ALFRESCO_AMHD().get_data()