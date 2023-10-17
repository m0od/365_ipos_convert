import glob
import os
import shutil
from datetime import datetime, timedelta
from os.path import dirname

import openpyxl



class YVES_ROCHER_AMHD(object):
    def __init__(self):
        self.ADAPTER_RETAILER = 'yves_rocher_amhd'
        self.ADAPTER_TOKEN = '62f862418ca872e0c98bfaa0133eef4f0d9c550ad454fe4a2a2e4d934414cf00'
        # self.ADAPTER_RETAILER = '695'
        # self.ADAPTER_TOKEN = 'cf0f760c3c11b65139beaecd6e0dd12f80bc34a177704ffc497d2bf816d1ac2d'

        self.FOLDER = 'yvesrocher_amhd'
        self.FULL_PATH = f'../home/{self.FOLDER}/'
        self.EXT = '*xlsx'
        self.DATA = None
        self.ORDERS = {}
        self.PMS = {}
        self.ODS = {}

    def scan_file(self):
        # print(os.getcwd())
        # print(os.path.getmtime(self.FULL_PATH))
        # print(for name in glob.glob('/home/geeks/Desktop/gfg/data.txt'):)
        files = glob.glob(self.FULL_PATH + self.EXT)
        # print(files)
        self.DATA = max(files, key=  os.path.getmtime)

        print(self.DATA)


    def get_data(self):
        self.scan_file()
        dataframe = openpyxl.load_workbook(self.DATA, data_only=True)

        sheet1 = dataframe['Danh sách đơn hàng']
        orders = {}
        # Iterate the loop to read the cell values
        for row in range(2, sheet1.max_row + 1):
            try:
                pm = []
                code = sheet1[row][0].value
                if code is None: continue
                code = str(code).strip()
                if len(code) == 0: continue
                code = code.replace("'", '')
                if code == '0': continue
                pur_date = sheet1[row][2].value
                print(code, pur_date)
                # pur_date = datetime.strptime(pur_date, '%Y-%m-%d %H:%M:%S')
                now = datetime.now() - timedelta(days=1)
                # if pur_date.replace(hour=0, minute=0) < now.replace(hour=0, minute=0, second=0, microsecond=0): continue
                pur_date = pur_date.strftime('%Y-%m-%d %H:%M:%S')
                try:
                    discount = float(sheet1[row][3].value)
                except:
                    discount = 0
                try:
                    total = float(sheet1[row][4].value) + discount
                except:
                    total = 0
                vat = float(sheet1[row][5].value)
                if orders.get(code) is None:
                    orders.update({code:{
                        'Code': code,
                        'Status': 2,
                        'PurchaseDate': pur_date,
                        'Total': total,
                        'TotalPayment': total,
                        'VAT': vat,
                        'Discount': abs(discount),
                    }})
                else:
                    orders[code].update({
                        'Code': code,
                        'Status': 2,
                        'PurchaseDate': pur_date,
                        'Total': orders[code]['Total'] + total,
                        'TotalPayment': orders[code]['Total'] + total,
                        'VAT': orders[code]['VAT'] + vat,
                        'Discount': orders[code]['Discount'] + abs(discount),
                    })
                # send = {
                #     'Code': code,
                #     'Status': 2,
                #     'PurchaseDate': pur_date,
                #     'Total': total,
                #     'TotalPayment': total,
                #     'VAT': vat,
                #     'Discount': 0,
                #     'OrderDetails': [],
                #     'PaymentMethods': pm
                # }
                # submit_order(retailer=self.ADAPTER_RETAILER, token=self.ADAPTER_TOKEN, data=send)
            except Exception as e:
                print(e)
                submit_error(retailer=self.ADAPTER_RETAILER, reason=str(e))

        sheet2 = dataframe['Phương thức thanh toán']
        pms = {}
        for row in range(2, sheet2.max_row + 1):
            code = sheet2[row][0].value
            # print(code)
            if code is None: continue
            code = str(code).strip()
            if len(code) == 0: continue
            code = code.replace("'", '')
            if code == '0': continue
            name = str(sheet2[row][1].value).strip().upper()
            if name == 'NONE' or name == 'TIỀN MẶT': name = 'CASH'
            # elif len(name) == 0:
            value = sheet2[row][2].value
            if pms.get(code) is None:
                pms.update({code: [{'Name': name, 'Value': value}]})
            else:
                pms[code].append({'Name': name, 'Value': value})
        for k, v in pms.items():
            if orders.get(k) is not None:
                orders[k].update({'PaymentMethods': v})
        sheet3 = dataframe['Chi tiết đơn hàng']
        ods = {}
        for row in range(2, sheet3.max_row + 1):
            code = sheet3[row][0].value
            # print(code)
            if code is None: continue
            code = str(code).strip()
            if len(code) == 0: continue
            code = code.replace("'", '')
            if code == '0': continue
            p_code = sheet3[row][1].value
            p_code = p_code.replace("'", '')
            # print(row, code, p_code)
            p_name = sheet3[row][2].value
            try:
                qty = float(sheet3[row][3].value)
            except:
                qty = 0
            try:
                price = sheet3[row][5].value
            except:
                price = 0
            if ods.get(code) is None:
                ods[code] = [{
                    'Code': p_code,
                    'Name': p_name,
                    'Price': price,
                    'Quantity': qty
                }]
            else:
                ods[code].append({
                    'Code': p_code,
                    'Name': p_name,
                    'Price': price,
                    'Quantity': qty
                })
        for k, v in ods.items():
            if orders.get(k) is not None:
                orders[k].update({'OrderDetails': v})
        for k, v in orders.items():
            if v.get('OrderDetails') is None:
                v.update({'OrderDetails': []})
            if v.get('PaymentMethods') is None:
                v.update({'PaymentMethods': [{'Name': 'CASH', 'Value': 0}]})
                # print(167, v)
            # print(v)
            submit_order(retailer=self.ADAPTER_RETAILER, token=self.ADAPTER_TOKEN, data=v)
        try:
            shutil.move(self.DATA, f'{self.FULL_PATH}bak')
        except:
            pass
if __name__:
    import sys

    PATH = dirname(dirname(__file__))
    # PATH = dirname(dirname(dirname(__file__)))
    # print(PATH)
    sys.path.append(PATH)
    from schedule.pos_api.adapter import submit_error, submit_order
    # YVES_ROCHER_AMHD().get_data()