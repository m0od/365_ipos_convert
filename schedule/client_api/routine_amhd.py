import glob
import os
from datetime import datetime, timedelta
from os.path import dirname

import openpyxl


class Routine(object):
    def __init__(self):
        self.ADAPTER_RETAILER = 'routine_aeonhd'
        self.ADAPTER_TOKEN = '9016a35bbb92076a998c777bb7cf3924ec1c211692fdc76314d277b0f496a41b'
        self.FOLDER = 'routine_amhd'
        self.FULL_PATH = f'../home/{self.FOLDER}/'
        self.EXT = '*xlsx'
        self.DATA = None
        self.ORDERS = {}
        self.PMS = {}
        self.ODS = {}

    def scan_file(self):
        # print(os.getcwd())
        # print(os.path.getmtime(self.FULL_PATH))
        # print(for name in glob.glob('/home/geeks/Desktop/gfg/data.txt'):)
        files = glob.glob(self.FULL_PATH + self.EXT)
        # print(files)
        self.DATA = max(files, key=os.path.getmtime)

        # print()

    def get_data(self):
        self.scan_file()
        dataframe = openpyxl.load_workbook(self.DATA, data_only=True)
        sheetOne = dataframe['Danh sách đơn hàng']
        for row in range(2, sheetOne.max_row + 1):
            code = sheetOne[row][0].value
            if code is None: continue
            if self.ORDERS.get(code) is None:
                self.ORDERS.update({str(code).strip(): {}})
            status = sheetOne[row][1].value
            pur_date = sheetOne[row][2].value
            pur_date = pur_date.strftime('%Y-%m-%d %H:%M:%S')
            now = datetime.now() - timedelta(days=1)
            if pur_date.day < now.day: continue
            discount = abs(sheetOne[row][3].value)
            total = sheetOne[row][4].value
            vat = sheetOne[row][5].value
            total_pm = sheetOne[row][6].value
            # print(type(pur_date))
            print(code, pur_date, type(total), type(discount), type(total_pm))
            if status == 2:
                self.ORDERS.update({code: {
                    'Code': str(code),
                    'Status': 2,
                    'PurchaseDate': pur_date,
                    'Total': total,
                    'TotalPayment': total_pm,
                    'VAT': vat,
                    'Discount': discount,
                }})

        sheetTwo = dataframe['Phương thức thanh toán']
        for row in range(2, sheetTwo.max_row + 1):
            code = sheetTwo[row][0].value
            if code is None: continue
            code = str(code).strip()
            name = str(sheetTwo[row][1].value).replace('TTTM', '').strip()
            value = sheetTwo[row][2].value
            print(code, name, value)

            if self.PMS.get(code) is None:
                self.PMS[code] = {
                    'PaymentMethods': [
                        {
                            'Name': name,
                            'Value': float(value)
                        }
                    ]
                }
            else:
                self.PMS[code]['PaymentMethods'].append(
                    {
                        'Name': name,
                        'Value': float(value)
                    }
                )

        for k, v in self.PMS.items():
            if self.ORDERS.get(k) is not None:
                self.ORDERS[k].update(v)

        sheetThree = dataframe['Chi tiết đơn hàng']

        # Iterate the loop to read the cell values
        self.ODS = {}
        for row in range(2, sheetThree.max_row + 1):
            code = sheetThree[row][0].value
            if code is None: continue
            code = str(code).strip()
            p_code = str(sheetThree[row][1].value)
            name = str(sheetThree[row][2].value)
            qty = sheetThree[row][3].value
            price = sheetThree[row][5].value
            price = price is not None and price or 0
            print('==>', code, p_code, name, float(price), int(qty))
            if self.ODS.get(code) is None:
                self.ODS[code] = {
                    'OrderDetails': [{
                        'Code': p_code,
                        'Name': name,
                        'Price': float(price),
                        'Quantity': int(qty)
                    }],
                }
            else:
                self.ODS[code]['OrderDetails'].append(
                    {
                        'Code': p_code,
                        'Name': name,
                        'Price': float(price),
                        'Quantity': int(qty)
                    }
                )

        for k, v in self.ODS.items():
            print(k, v)
            if self.ORDERS.get(k) is not None:
                self.ORDERS[k].update(v)
                print(self.ORDERS[k])
            else:
                print('unk')

        for k, v in self.ORDERS.items():
            print(v)
            submit_order(retailer=self.ADAPTER_RETAILER, token=self.ADAPTER_TOKEN, data=v)


if __name__:
    import sys

    PATH = dirname(dirname(__file__))
    print(PATH)
    sys.path.append(PATH)
    from schedule.pos_api.adapter import submit_error, submit_order
    # a = Routine()
    # a.get_data()
