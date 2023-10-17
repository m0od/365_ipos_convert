import glob
import os
import shutil

import xlrd
from datetime import datetime, timedelta
from os.path import dirname

import openpyxl



class VERA_JOCKEY_AMHD(object):
    def __init__(self):
        self.ADAPTER_RETAILER = 'verajockey_amhd'
        self.ADAPTER_TOKEN = 'f6163a245cd2ac70fc6aa477eb3f9af4162fdcf48b979048ec94e2b9edf95712'
        # self.ADAPTER_RETAILER = '695'
        # self.ADAPTER_TOKEN = 'cf0f760c3c11b65139beaecd6e0dd12f80bc34a177704ffc497d2bf816d1ac2d'

        self.FOLDER = 'verajockey_amhd'
        self.FULL_PATH = f'../home/{self.FOLDER}/'
        self.EXT = '*xls*'
        self.DATA = None
        self.ORDERS = {}
        self.PMS = {}
        self.ODS = {}

    def scan_file(self):
        # print(os.getcwd())
        # print(os.path.getmtime(self.FULL_PATH))
        # print(for name in glob.glob('/home/geeks/Desktop/gfg/data.txt'):)
        files = glob.glob(self.FULL_PATH + self.EXT)
        # print(files)
        self.DATA = max(files, key=  os.path.getmtime)
        # if self.DATA.endswith('.xls'):
        #     os.rename(self.DATA, self.DATA + 'x')
        #     self.DATA += 'x'
        print(self.DATA)


    def get_data(self):
        self.scan_file()
        if not self.DATA.endswith('.xls'):
            dataframe = openpyxl.load_workbook(self.DATA, data_only=True)
        else:
            dataframe = xlrd.open_workbook(self.DATA)
        sheet1 = dataframe['Danh sách đơn hàng']
        orders = {}
        if not self.DATA.endswith('.xls'):
            nRows = sheet1.max_row + 1
        else:
            nRows = sheet1.nrows
        # Iterate the loop to read the cell values
        for row in range(2, nRows):
            try:
                pm = []
                code = sheet1[row][0].value
                if code is None: continue
                code = str(code).strip()
                code = code.replace("'", '').strip()
                if len(code) == 0: continue
                pur_date = sheet1[row][2].value
                pur_date = datetime.strptime(pur_date, '%d/%m/%Y %H:%M %p')
                now = datetime.now() - timedelta(days=1)
                # if pur_date.replace(hour=0, minute=0) < now.replace(hour=0, minute=0, second=0, microsecond=0): continue
                pur_date = pur_date.strftime('%Y-%m-%d %H:%M:%S')
                try:
                    discount = float(sheet1[row][3].value)
                except:
                    discount = 0
                try:
                    total = float(sheet1[row][4].value)
                except:
                    total = 0
                vat = sheet1[row][5].value
                if orders.get(code) is None:
                    orders.update({code:{
                        'Code': code,
                        'Status': 2,
                        'PurchaseDate': pur_date,
                        'Total': total,
                        'TotalPayment': total,
                        'VAT': vat,
                        'Discount': discount,
                    }})
                else:
                    orders[code].update({
                        'Code': code,
                        'Status': 2,
                        'PurchaseDate': pur_date,
                        'Total': orders[code]['Total'] + total,
                        'TotalPayment': orders[code]['Total'] + total,
                        'VAT': orders[code]['VAT'] + vat,
                        'Discount': orders[code]['Discount'] + discount,
                    })
            except Exception as e:
                print(e)
                submit_error(retailer=self.ADAPTER_RETAILER, reason=str(e))
        sheet2 = dataframe['Phương thức thanh toán']
        pms = {}
        if not self.DATA.endswith('.xls'):
            nRows = sheet2.max_row + 1
        else:
            nRows = sheet2.nrows
        for row in range(2, nRows):
            code = sheet2[row][0].value
            if code is None: continue
            code = str(code).strip()
            code = code.replace("'", '').strip()
            # print(code)
            if len(code) == 0: continue
            # print(code)
            name = str(sheet2[row][2].value).strip().upper()
            if name == 'NONE' or name == 'TIỀN MẶT': name = 'CASH'
            # elif len(name) == 0:
            value = sheet2[row][1].value
            if pms.get(code) is None:
                pms.update({code: [{'Name': name, 'Value': value}]})
            else:
                pms[code].append({'Name': name, 'Value': value})
        for k, v in pms.items():
            if orders.get(k) is not None:
                orders[k].update({'PaymentMethods': v})
        sheet3 = dataframe['Chi tiết đơn hàng']
        ods = {}
        if not self.DATA.endswith('.xls'):
            nRows = sheet3.max_row + 1
        else:
            nRows = sheet3.nrows
        for row in range(2, nRows):
            code = sheet3[row][0].value
            # print(code)
            if code is None: continue
            code = str(code).strip()
            code = code.replace("'", '').strip()
            if len(code) == 0: continue
            p_code = sheet3[row][1].value
            p_code = p_code.replace("'", '')
            # print(row, code, p_code)
            p_name = sheet3[row][2].value
            try:
                qty = float(sheet3[row][3].value)
            except:
                qty = 0
            try:
                price = sheet3[row][5].value
            except:
                price = 0
            if ods.get(code) is None:
                ods[code] = [{
                    'Code': p_code,
                    'Name': p_name,
                    'Price': price,
                    'Quantity': qty
                }]
            else:
                ods[code].append({
                    'Code': p_code,
                    'Name': p_name,
                    'Price': price,
                    'Quantity': qty
                })
        for k, v in ods.items():
            if orders.get(k) is not None:
                orders[k].update({'OrderDetails': v})
        for k, v in orders.items():
            # print(v)
            if v.get('OrderDetails') is None:
                v.update({'OrderDetails': []})
            if v.get('PaymentMethods') is None:
                print(v)
            submit_order(retailer=self.ADAPTER_RETAILER, token=self.ADAPTER_TOKEN, data=v)
        try:
            shutil.move(self.DATA, f"{self.FULL_PATH}bak")
        except:
            pass
if __name__:
    import sys

    PATH = dirname(dirname(__file__))
    # PATH = dirname(dirname(dirname(__file__)))
    # print(PATH)
    sys.path.append(PATH)
    from schedule.pos_api.adapter import submit_error, submit_order
    # VERA_JOCKEY_AMHD().get_data()