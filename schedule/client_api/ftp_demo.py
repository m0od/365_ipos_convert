from datetime import datetime

import openpyxl
from openpyxl.formula.translate import Translator

from schedule.pos_api.adapter import submit_order

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("110_ROUTINE_IMPORT 10.06 (1).xlsx", data_only=True)

# Read SheetOne
sheetOne = dataframe['Danh sách đơn hàng']
orders = {}
# Iterate the loop to read the cell values
for row in range(2, sheetOne.max_row):
    # print(sheetOne[row][0].value)
    code = sheetOne[row][0].value
    if code is None: break
    code = str(code)
    status = str(sheetOne[row][1].value)
    pur_date = str(sheetOne[row][2].value)
    # print(pur_date)
    pur_date = datetime.strptime(pur_date, '%Y-%m-%d %H:%M:%S')
    pur_date = pur_date.strftime('%Y-%m-%d %H:%M:%S')
    discount = str(sheetOne[row][3].value)
    total = str(sheetOne[row][4].value)
    vat = str(sheetOne[row][5].value)
    # print(code, status, pur_date, discount, total, vat)
    if orders.get(code) is None:
        orders[code] = {
            'Code': code,
            'Status': 2,
            'PurchaseDate': pur_date,
            'Total': total,
            'TotalPayment': total,
            'VAT': vat,
            'Discount': abs(float(discount)),

        }
    # else:
    #     orders[code].update()
    # for col in sheetOne.iter_cols(1, sheetOne.max_column):
    #     if col[row].value is None: break
    #     print(col[row].value)
    # print('-'*5)

# Read SheetTwo
sheetTwo = dataframe['Phương thức thanh toán']

# Iterate the loop to read the cell values
pms = {}
for row in range(2, sheetTwo.max_row):
    code = sheetTwo[row][0].value
    if code is None: break
    code = str(code)
    name = str(sheetTwo[row][1].value)
    value = str(sheetTwo[row][2].value)
    # print(name, value)
    if pms.get(code) is None:
        pms[code] = {
            'PaymentMethods': [
                {
                    'Name': name,
                    'Value': float(value)
                }
            ]
        }
    else:
        pms[code]['PaymentMethods'].append(
            {
                'Name': name,
                'Value': float(value)
            }
        )
for k, v in pms.items():
    if orders.get(k) is not None:
        orders[k].update(v)
#         print(orders[k])

sheetThree = dataframe['Chi tiết đơn hàng']

# Iterate the loop to read the cell values
ods = {}
for row in range(2, sheetThree.max_row):
    code = sheetThree[row][0].value
    print(code)
    if code is None: continue
    code = str(code).strip()
    p_code = str(sheetThree[row][1].value)
    name = str(sheetThree[row][2].value)
    qty = str(sheetThree[row][3].value)
    price = str(sheetThree[row][5].value)
    print('==>',code, p_code, name, float(price), int(qty))
    if ods.get(code) is None:
        ods[code] = {
            'OrderDetails': [{
                'Code': p_code,
                'Name': name,
                'Price': float(price),
                'Quantity': int(qty)
            }],
        }
    else:
        ods[code]['OrderDetails'].append(
            {
                'Code': p_code,
                'Name': name,
                'Price': float(price),
                'Quantity': int(qty)
            }
        )

for k, v in ods.items():
    print(k, v)
    if orders.get(k) is not None:
        orders[k].update(v)
        print(orders[k])
    else:
        print('unk')

for k, v in orders.items():
    print(v)
    if v.get('PaymentMethods') is None:
        v['PaymentMethods'] = [{'Name': 'CASH', 'Value': 0}]
    if v.get('OrderDetails') is None:
        v['OrderDetails'] = []
    submit_order(retailer='elisetest', token='f26ac701e8018f4cd1259161ec2b64881c0b4b05009add0183f0c9db9535ad56', data=v)