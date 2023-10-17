import glob
import os
import shutil
import smtplib
from datetime import datetime, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from os.path import dirname

import openpyxl


class PNJ_AMHD(object):
    def __init__(self):
        self.ADAPTER_RETAILER = 'pnj{}_aeonhd'
        self.ADAPTER_TOKEN = '22e07b3b942190b5b91eaf88b8c7937741fcbbf9932def888c1aad9aa72fcba5'
        self.FOLDER = 'pnj_amhd'
        self.FULL_PATH = f'../home/{self.FOLDER}/'
        self.EXT = '*xlsx'
        self.DATA = None
        self.ORDERS = {}
        self.PMS = {}
        self.ODS = {}

    def scan_file(self):
        # print(os.getcwd())
        # print(os.path.getmtime(self.FULL_PATH))
        # print(for name in glob.glob('/home/geeks/Desktop/gfg/data.txt'):)
        files = glob.glob(self.FULL_PATH + self.EXT)
        # print(files)
        self.DATA = max(files, key=os.path.getmtime)

        print(self.DATA)

    def get_data(self):
        self.scan_file()
        dataframe = openpyxl.load_workbook(self.DATA, data_only=True)

        sheetOne = dataframe['Danh sách đơn hàng']
        # print(sheetOne)
        orders = {}
        # Iterate the loop to read the cell values
        for row in range(2, sheetOne.max_row + 1):
            # print(sheetOne[row][0].value)
            code = sheetOne[row][0].value
            if code is None: break
            code = str(code)
            status = str(sheetOne[row][1].value).strip()
            pur_date = str(sheetOne[row][2].value).strip()
            pur_date = datetime.strptime(pur_date, '%d%m%Y%H%M')
            now = datetime.now() - timedelta(days=1)
            if pur_date.replace(hour=0, minute=0) < now.replace(hour=0, minute=0, second=0, microsecond=0): continue
            pur_date = pur_date.strftime('%Y-%m-%d %H:%M:%S')
            discount = str(sheetOne[row][3].value)
            total = str(sheetOne[row][4].value)
            vat = str(sheetOne[row][5].value)
            branch = sheetOne[row][6].value
            if orders.get(code) is None:
                orders[code] = {
                    'Code': code,
                    'Branch': branch,
                    'Status': int(status),
                    'PurchaseDate': pur_date,
                    'Total': float(total),
                    'TotalPayment': float(total),
                    'VAT': float(vat),
                    'Discount': abs(float(discount)),

                }
        sheetTwo = dataframe['Phương thức thanh toán']
        pms = {}
        for row in range(2, sheetTwo.max_row + 1):
            code = sheetTwo[row][0].value
            if code is None: break
            code = str(code)
            name = str(sheetTwo[row][1].value)
            value = str(sheetTwo[row][2].value)
            if pms.get(code) is None:
                pms[code] = {
                    'PaymentMethods': [
                        {
                            'Name': name,
                            'Value': float(value)
                        }
                    ]
                }
            else:
                pms[code]['PaymentMethods'].append(
                    {
                        'Name': name,
                        'Value': float(value)
                    }
                )
        # print(pms)
        for k, v in pms.items():
            if orders.get(k) is not None:
                orders[k].update(v)
                # print(orders[k])
        #
        sheetThree = dataframe['Chi tiết đơn hàng']

        # Iterate the loop to read the cell values
        ods = {}
        for row in range(2, sheetThree.max_row + 1):
            code = sheetThree[row][0].value
            # print(code)
            if code is None: continue
            code = str(code).strip()
            p_code = str(sheetThree[row][1].value)
            name = str(sheetThree[row][2].value)
            qty = str(sheetThree[row][3].value)
            price = str(sheetThree[row][4].value)
            # print('==>', code, p_code, name, float(price), int(qty))
            if ods.get(code) is None:
                ods[code] = {
                    'OrderDetails': [{
                        'Code': p_code,
                        'Name': name,
                        'Price': float(price),
                        'Quantity': float(qty)
                    }],
                }
            else:
                ods[code]['OrderDetails'].append(
                    {
                        'Code': p_code,
                        'Name': name,
                        'Price': float(price),
                        'Quantity': float(qty)
                    }
                )
        for k, v in ods.items():
            # print(k, v)
            if orders.get(k) is not None:
                orders[k].update(v)
                # print(orders[k])

        for k, v in orders.items():
            if v.get('PaymentMethods') is None:
                v['PaymentMethods'] = [{'Name': 'CASH', 'Value': v['Total']}]
            if v.get('OrderDetails') is None:
                v['OrderDetails'] = []
            retailer = self.ADAPTER_RETAILER.format(str(int(v['Branch'])))
            v.pop('Branch')
            if v['Status'] == 2:
                print(v)
                submit_order(retailer=retailer, token=self.ADAPTER_TOKEN, data=v)
            elif v['Status'] == 1:
                send = v.copy()
                send.update({
                    # 'Code': f'VAT_{send["Code"]}',
                    'ReturnDate': send['PurchaseDate'],
                    'Total': abs(send['Total']),
                    'TotalPayment': abs(send['TotalPayment']),
                    'Status': 0,
                    'ReturnDetails': send['OrderDetails'],
                    'PaymentMethods': send['PaymentMethods']
                })
                send.pop('PurchaseDate')
                send.pop('OrderDetails')
                # print(send)
                submit_order(retailer=retailer, token=self.ADAPTER_TOKEN, data=send)
        if len(orders.items()) == 0:
            now = datetime.now() - timedelta(days=1)
            port = 465  # For SSL
            password = 'abqqzkkrgftlodny'
            smtp_server = "smtp.gmail.com"
            sender_email = "tungpt@pos365.vn"  # Enter your address
            to_email = 'vuong.tt01@pnj.com.vn'  # Enter receiver address
            cc_email = 'tungpt@pos365.vn'
            message = MIMEMultipart("alternative")
            message["Subject"] = f'Report 0 orders PNJ-AMHD {now.strftime("%Y-%m-%d")}'
            message["From"] = 'tungpt@pos365.vn'
            message["To"] = to_email
            toAddr = [to_email]
            while True:
                try:
                    server = smtplib.SMTP_SSL("smtp.gmail.com", 465)
                    server.login(sender_email, password)
                    server.sendmail(sender_email, toAddr, message.as_string())
                    break
                except Exception as e:
                    print(e)
                    pass
        try:
            shutil.move(self.DATA, f"{self.FULL_PATH}bak")
        except:
            pass


if __name__:
    import sys

    PATH = dirname(dirname(__file__))
    # PATH = dirname(dirname(dirname(__file__)))
    # print(PATH)
    sys.path.append(PATH)
    from schedule.pos_api.adapter import submit_error, submit_order
    # PNJ_AMHD().get_data()
