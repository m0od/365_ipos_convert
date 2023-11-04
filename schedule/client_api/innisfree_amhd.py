import glob
import os
import shutil
import openpyxl
from datetime import datetime, timedelta
from os.path import dirname


class INNIS_FREE_AMHD(object):
    def __init__(self):
        self.ADAPTER_RETAILER = 'innisfree_amhd'
        self.ADAPTER_TOKEN = '9efbf2b6aa6337f497c9a060b0fc2e658f27d06af0b514de3d50ba552af7f00a'
        # self.ADAPTER_RETAILER = '695'
        # self.ADAPTER_TOKEN = 'cf0f760c3c11b65139beaecd6e0dd12f80bc34a177704ffc497d2bf816d1ac2d'

        self.FOLDER = 'innisfree_amhd'
        self.FULL_PATH = f'../home/{self.FOLDER}/'
        self.EXT = '*xlsx'
        self.DATA = None
        self.ORDERS = {}
        self.PMS = {}
        self.ODS = {}

    def scan_file(self):
        files = glob.glob(self.FULL_PATH + self.EXT)
        self.DATA = max(files, key=os.path.getmtime)
        print(self.DATA)

    def get_data(self):
        self.scan_file()
        dataframe = openpyxl.load_workbook(self.DATA, data_only=True)
        sheet = dataframe[dataframe.sheetnames[0]]
        orders = {}
        nRows = sheet.max_row + 1
        for row in range(3, nRows - 1):
            try:
                code = sheet[row][1].value
                pur_date = f'{sheet[row][2].value} {sheet[row][3].value}'
                pur_date = datetime.strptime(pur_date, '%Y-%m-%d %H:%M:%S')
                pur_date = pur_date.strftime('%Y-%m-%d %H:%M:%S')
                E = float(sheet[row][4].value)
                F = float(sheet[row][5].value)
                P = float(sheet[row][15].value)
                total = E - F - P
                discount = F + P
                vat = float(sheet[row][7].value)
                cash = float(sheet[row][8].value)
                card = float(sheet[row][9].value)
                card += float(sheet[row][10].value)
                card += float(sheet[row][11].value)
                card += float(sheet[row][12].value)
                card += float(sheet[row][13].value)
                card += float(sheet[row][14].value)
                voucher = float(sheet[row][17].value)
                gift_card = float(sheet[row][18].value)
                moca = float(sheet[row][19].value)
                now = float(sheet[row][20].value)
                air_pay = float(sheet[row][21].value)
                got_it = float(sheet[row][22].value)
                vn_pay = float(sheet[row][23].value)
                ur_box = float(sheet[row][24].value)
                gift_pop = float(sheet[row][25].value)
                cod = float(sheet[row][26].value)
                atome = float(sheet[row][27].value)
                momo = float(sheet[row][28].value)
                pm = {
                    'CASH': cash,
                    'THẺ': card,
                    'VOUCHER': voucher,
                    'GIFT CARD': gift_card,
                    'MOCA': moca,
                    'NOW': now,
                    'AIRPAY': air_pay,
                    'GOT IT': got_it,
                    'VNPAY': vn_pay,
                    'URBOX': ur_box,
                    'GIFT POP': gift_pop,
                    'COD': cod,
                    'ATOME': atome,
                    'MOMO': momo
                }
                keys = list(pm.keys())
                for _ in range(len(keys)):
                    if pm[keys[_]] == 0:
                        pm.pop(keys[_])
                if pm == {}:
                    pm = {'CASH': 0}
                pms = []
                for k, v in pm.items():
                    pms.append({'Name': k, 'Value': v})
                # print(pms)

                self.ORDERS.update({code: {
                    'Code': code,
                    'Status': 2,
                    'PurchaseDate': pur_date,
                    'Total': total,
                    'TotalPayment': total,
                    'VAT': vat,
                    'Discount': discount,
                    'PaymentMethods':pms,
                    'OrderDetails': []
                }})

            except Exception as e:
                print(e)
                submit_error(retailer=self.ADAPTER_RETAILER, reason=str(e))
        for k, v in self.ORDERS.items():
            submit_order(retailer=self.ADAPTER_RETAILER, token=self.ADAPTER_TOKEN, data=v)
        try:
            shutil.move(self.DATA, f"{self.FULL_PATH}bak")
        except:
            pass


if __name__:
    import sys

    PATH = dirname(dirname(__file__))
    # PATH = dirname(dirname(dirname(__file__)))
    # print(PATH)
    sys.path.append(PATH)
    from schedule.pos_api.adapter import submit_error, submit_order

    # INNIS_FREE_AMHD().get_data()
